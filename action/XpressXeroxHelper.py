import os
import shutil
from pathlib import Path
from Xpressxerox.settings import MEDIA_ROOT
from django.utils import timezone
from django.core.files.storage import FileSystemStorage
from PyPDF2 import PdfFileReader
from openpyxl import Workbook, load_workbook
import zipfile

""""
BELOW FUNCTIONS ARE FOR HELPING PURPOSE .
LIST OF ALL FUNCTIONS

1. getFileList
2. checkExtention
3. uploadFiles
"""

def getFileList(user):
    """
    user has 3 sub folder B&w , back2back and color.
    this will tell sequential files in each folder

    """
    fileList = []
    for i in "123":
        destination = Path.joinpath(MEDIA_ROOT, user)
        destination = Path.joinpath(destination, i)
        files = sorted(os.listdir(destination))
        #print(destination,"\n",files)
        fileList.append(files)
    return fileList

def checkExtention(files):
    """
    This function will check extention of each file.

    """
    for file in files:
        if file.name.lower().split(".")[-1] not in ['pdf', "jpeg", "png", "jpg"]:
            return file.name + " is not accepted. We accept only 'pdf', 'jpeg', 'png', 'jpg' file format."
    return "Files Verified"

def isPdf(file):
    if file.lower().split(".")[-1] == "pdf":
        return True
    else:
        return False

def isEncrptedPdf(files):
    for file in files:
        if PdfFileReader(file).isEncrypted:
            return file.name + " is Password Protected File"
        else:
            pass
    return "Files are not encrypted"

def uploadFiles(files, user, option):
    """
    This function work on option argument option will tell where to stor file
    if option is 1 file will store in sub folder 1 of user which is for Print per page B&W.
    option 2 tell sub folder 2 of user which for Print BacktoBack on page with B&W.
    option 3 for sub folder 3 of user which tells print per page is color.

    """

    destination = Path.joinpath(MEDIA_ROOT, user)
    current_files =[]
    current_bill = 0
    if option == "1":
        destination = Path.joinpath(destination, "trash", "1")
    elif option == "2":
        destination = Path.joinpath(destination, "trash", "2")
    elif option == "3":
        destination = Path.joinpath(destination,"trash", "3")

    for file in files:
        fs = FileSystemStorage(location=destination)
        filename = fs.save(file.name, file)
        filename = os.path.join(destination, filename)
        current_files.append(filename)
        if isPdf(file.name):
            pages = pageCounter(filename)
        else:
            pages = 1

        if option == "1":
            current_bill += pages

        elif option == "2":
            if pages > 59:
                current_bill = current_bill + ((pages%2 + pages//2)*1)
            elif pages > 40:
                current_bill = current_bill + ((pages%2 + pages//2)*1.5)
            else:
                current_bill = current_bill + ((pages%2 + pages//2)*2)

        elif option == "3":
            current_bill += (pages * 10)
    print("current_files:- ", current_files, "\ncurrent_bill:- ", current_bill)
    return current_files, current_bill

def paymentStatus(session, act = 0):
    """
    check the keys current_bill n current_files are present then remove the files if transaction not completed
    """
    if act == 1:
        print("--------tran succesful----------")
        for file in session["current_files"]:
            des = file.split("trash")
            print(des)
            des = des[0] + des[1][1:]
            print("Moving from ---->", file)
            print("moved to ---->", des)
            shutil.move(src= file, dst= des)

    else:
        for file in session["current_files"]:
            print("file removed -->", file)
            os.remove(file)




def uniqueId(username):
    """
    Using username and timezone will return UniqueId username+ now
    """
    username = username
    now = timezone.now()
    now = now.strftime('-%Y%m%d-%H%M%S')
    return str(username+now)

def pageCounter(file):
    """
        parameters:- file - path of file
        count number of pages in file
    """
    temp = open(file, "rb")
    pdf = PdfFileReader(temp)
    pages = pdf.getNumPages()
    temp.close()
    print("document ", file, "pages ", pages)

    return int(pages)


def createBill(bill, users, record):
    """
        parameters:- bill - path of bill txt files
                    users - list of userss
                    record - dict of users with pages
    """
    file = open(bill, 'w')
    header = "USERNAME \t PAGES \t REMARK\n"
    file.write(header)
    for user in users:
        entry = "{} \t {} \t Yes/No\n".format(user, record[user])
        file.write(entry)
    file.close()

    return None

def reformatBill(bill):
    """
    parameters:- bill path which created by createBill()
    Returns:- Reformated bill with spaces
    """
    with open(bill) as f:
        datatable = [line.split() for line in f.read().splitlines()]
    header = datatable[0]
    datatable = datatable[1:]
    widths = [max(len(value) for value in col)
              for col in zip(*(datatable + [header]))]
    format_spec = '{:{widths[0]}}  {:>{widths[1]}}  {:>{widths[2]}}'
    header = format_spec.format(*header, widths=widths)
    file = open(bill, 'w')
    file.write(header)
    for fields in datatable:
        file.write("\n")
        entry = format_spec.format(*fields, widths=widths)
        file.write(entry)
    file.close()
    return None

# def createBillExcel(bill):
#     """
#         parameters:- bill - path of bill xcel files
#         create empty document
#     """
#     wb = Workbook()
#     wb.save(bill)
#     wb.close()
#     print("--------Bill file created--------")
#     return None

# def updatedBill(bill, users, record):
#     """
#         parameters:- bill - path of bill xcel files, users :- sorted list of user send doc to print,
#         record dict has record of users and how many pages they want to print
#     """
#     wb = load_workbook(bill)
#     sheet = wb["Sheet"]
#     sheet.cell(row=1, column=1).value = "UserName"
#     sheet.cell(row=1, column=2).value = "Pages"
#     r = 2
#     for entry in users:
#         sheet.cell(row=r, column=1).value = entry
#         sheet.cell(row=r, column=2).value = record[entry]
#         r += 1

#     wb.save(bill)
#     wb.close()
#     print("--------Bill file Updated--------")
#     return None

def createZip(zip_name, src):
    """
        parameter:- zip file name , folder path
        This will create zip file
    """
    zipf = zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED)
    for file in src:
        zipf.write(file)
    zipf.close()
    print("--------zip file created --------")
    return None


def deleteFiles(src):
    """
        parameter:- folder path with files
        This will delete files from folder
    """
    for file in src:
        os.remove(file)
    print("--------deleted files-------")
    return None

def clearTrash(user):
    destination = Path.joinpath(MEDIA_ROOT, user, "trash")
    files = [ f for f in os.listdir(destination) if os.path.isfile(Path.joinpath(destination, f)) ]
    if files:
        print("---clearTrash---", files)
        for file in files:
            file = Path.joinpath(destination, file)
            os.remove(file)
    for folder in "123":
        trash = Path.joinpath(destination, folder)
        files = [f for f in os.listdir(trash) if os.path.isfile(Path.joinpath(trash, f))]
        if files:
            print("---clearTrash---", files)
            for file in files:
                file = Path.joinpath(trash, file)
                os.remove(file)

def saveDocx(files, user):
    destination = Path.joinpath(MEDIA_ROOT, user, "trash")
    current_doc_files = list()
    for file in files:
        fs = FileSystemStorage(location=destination)
        filename = fs.save(file.name, file)
        filename = os.path.join(destination, filename)
        current_doc_files.append(filename)
    print("doc files---->", current_doc_files)
    return current_doc_files, destination

def convToPdf(current_doc_files, user):
    current_pdf_files = list()
    for file in current_doc_files:
        cmd = "abiword --to=pdf '{}'".format(file)
        print("----cmd----", cmd)
        try:
            os.system(cmd)
            pdf_file_name = os.path.splitext(file)[0] + '.pdf'
            print("converted ", pdf_file_name)
            current_pdf_files.append(pdf_file_name)
        except Exception as ex:
            print("--------Exception--------")
            print(ex)
    return current_pdf_files


def checkDoc(files):
    for file in files:
        if file.name.lower().split(".")[-1] not in ['docx', "doc"]:
            return file.name + " is not Docx or Doc type file."
    return "Files Verified"

